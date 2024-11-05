import csv
from openpyxl import Workbook
from datetime import datetime

# Функція для обчислення віку
def calculate_age(birth_date):
    return datetime.now().year - birth_date.year

# Читання CSV файлу та розподіл за віковими категоріями
try:
    with open("employees.csv", mode="r", encoding="utf-8") as file:
        reader = csv.reader(file)
        next(reader)  # Пропустити заголовок
        data = list(reader)
    
    workbook = Workbook()
    sheets = {
        "all": workbook.active,
        "younger_18": workbook.create_sheet("younger_18"),
        "18-45": workbook.create_sheet("18-45"),
        "45-70": workbook.create_sheet("45-70"),
        "older_70": workbook.create_sheet("older_70"),
    }
    sheets["all"].title = "all"
    headers = ["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"]
    for sheet in sheets.values():
        sheet.append(headers)

    for index, row in enumerate(data, start=1):
        birth_date = datetime.strptime(row[4], "%Y-%m-%d")
        age = calculate_age(birth_date)
        new_row = [index] + row[:3] + [row[4], age]

        sheets["all"].append(new_row)
        if age < 18:
            sheets["younger_18"].append(new_row)
        elif 18 <= age <= 45:
            sheets["18-45"].append(new_row)
        elif 45 < age <= 70:
            sheets["45-70"].append(new_row)
        else:
            sheets["older_70"].append(new_row)

    workbook.save("employees.xlsx")
    print("Ok")
except FileNotFoundError:
    print("Помилка: Файл employees.csv не знайдено.")
except Exception as e:
    print("Помилка при створенні файлу XLSX:", e)
